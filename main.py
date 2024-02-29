import openpyxl
from datetime import datetime, timedelta
from openpyxl import styles

meses = {
    'ENERO': 'January',
    'FEBRERO': 'February',
    'MARZO': 'March',
    'ABRIL': 'April',
    'MAYO': 'May',
    'JUNIO': 'June',
    'JULIO': 'July',
    'AGOSTO': 'August',
    'SEPTIEMBRE': 'September',
    'OCTUBRE': 'October',
    'NOVIEMBRE': 'November',
    'DICIEMBRE': 'December',
}

dias = {
    'Monday': 'lunes',
    'Tuesday': 'martes',
    'Wednesday': 'miércoles',
    'Thursday': 'jueves',
    'Friday': 'viernes',
    'Saturday': 'sábado',
    'Sunday': 'domingo',
}

# INGRESAR MES Y AÑO
mes = input("Ingrese el nombre del mes: ")
mes = mes.upper()

# VERIFICAR QUE EL NOMBRE DEL MES SEA VALIDO
if mes not in meses:
    print("Nombre de mes no válido. Por favor, ingrese un nombre de mes en español válido.")
    exit()

# INGRESAR AÑO
try:
    año = int(input("Ingrese el año: "))
except ValueError:
    print("Por favor, ingrese un año válido.")
    exit()

# OBTENER DATOS DEL MES
mes_ingles = meses[mes]
mes_num = datetime.now().replace(month=list(meses.values()).index(mes_ingles) + 1, day=1, year=año).month
primer_dia_mes = datetime.now().replace(month=list(meses.values()).index(mes_ingles) + 1, day=1, year=año).weekday()
ultimo_dia_mes = (datetime.now().replace(month=list(meses.values()).index(mes_ingles) + 1, day=1, year=año) + timedelta(days=31)).replace(day=1) - timedelta(days=1)

# CREAR EXCEL
nuevo_excel = openpyxl.Workbook()
nuevo_excel.remove(nuevo_excel.active)

plantilla = openpyxl.load_workbook('PLANTILLA.xlsx')
hoja_tipo = plantilla['TIPO']

for dia in range(1, ultimo_dia_mes.day + 1):
    dia_actual = datetime.now().replace(month=mes_num, day=dia, year=año)
    if dia_actual.weekday() == 6:
        continue   

    dia_semana = dias[dia_actual.strftime('%A')]
    nueva_hoja = nuevo_excel.create_sheet(title=f'{dia_semana}, {dia}')

    # COPIAR CELDAS Y ESTILO DE LA HOJA DE PLANTILLA
    for row_index, row in enumerate(hoja_tipo.iter_rows(), start=1):
        for col_index, cell in enumerate(row, start=1):
            nueva_celda = nueva_hoja.cell(row=row_index, column=col_index, value=cell.value)

            nueva_celda.font = styles.Font(bold=cell.font.bold)
            nueva_celda.fill = styles.PatternFill(start_color=cell.fill.start_color, end_color=cell.fill.end_color, fill_type=cell.fill.fill_type)
            nueva_celda.alignment = styles.Alignment(horizontal=cell.alignment.horizontal, vertical=cell.alignment.vertical, wrap_text=cell.alignment.wrap_text)

            nueva_celda.border = styles.Border(
                left=styles.Side(style=cell.border.left.style, color=cell.border.left.color),
                right=styles.Side(style=cell.border.right.style, color=cell.border.right.color),
                top=styles.Side(style=cell.border.top.style, color=cell.border.top.color),
                bottom=styles.Side(style=cell.border.bottom.style, color=cell.border.bottom.color),
            )

    for col_index, column_dimension in hoja_tipo.column_dimensions.items():
        nueva_hoja.column_dimensions[col_index].width = column_dimension.width

    nueva_hoja['F4'] = f'{dia_semana}, {dia}'

# GUARDAR
nuevo_excel.save(f'{mes} {año}.xlsx')
